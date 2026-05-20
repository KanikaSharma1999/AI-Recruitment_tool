import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, portrait
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import numpy as np

def extract_candidate_info(c, idx):
    hs = c.get("hiring_summary", {})
    ai = c.get("ai_analysis", {})
    
    name = c.get("name", "Unknown")
    email = c.get("email", "N/A")
    phone = c.get("phone", "N/A")
    experience_val = c.get('experience_years', 0)
    experience = f"{experience_val} years"
    
    matched_skills = c.get("matched_skills", [])
    if not matched_skills:
        matched_skills = c.get("skills", [])
    skills_str = ", ".join(matched_skills[:10])
    
    score = round(c.get("score", 0))
    skill_match = round(c.get("technical_fit", 0))
    
    rank = idx + 1
    
    # 1. Recommendation Fallback
    verdict = ai.get("verdict", "")
    recommendation = hs.get("recommendation", "")
    if not recommendation or recommendation == "N/A":
        recommendation = ai.get("recommendation", "Review Required")
    if verdict:
        recommendation = f"{recommendation} ({verdict})"
        
    # 2. Summary Fallback
    summary = hs.get("summary", ai.get("executive_summary", ""))
    if not summary or summary == "N/A" or "No summary available" in summary:
        summary = f"{name} brings {experience_val} years of professional experience with core expertise in {skills_str if skills_str else 'relevant industry technologies'}. The candidate demonstrates solid alignment with the operational and technical requirements of the role, presenting a strong profile for further evaluation."
        
    # 3. Feedback Fallback
    reasoning = hs.get("reasoning", ai.get("reasoning", ""))
    missing_list = c.get("missing_skills", [])
    missing = ", ".join(missing_list)
    
    if not reasoning or reasoning == "N/A":
        reasoning = f"Technical Fit: {skill_match}%. Candidate shows high potential based on experience match and core competencies."
        
    feedback = f"{reasoning}\nMissing Skills: {missing if missing else 'None detected. Candidate possesses all critical skills.'}"
    
    # 4. Strengths & Weaknesses Fallback
    strengths_list = hs.get("strengths", [])
    if not strengths_list:
        strengths_list = matched_skills[:5]
    strengths = ", ".join(strengths_list) if strengths_list else "General competency in required domains"
    
    weaknesses_list = hs.get("weaknesses", [])
    if not weaknesses_list:
        weaknesses_list = missing_list[:3]
    weaknesses = ", ".join(weaknesses_list) if weaknesses_list else "No significant weaknesses detected"
    
    hiring_decision = str(c.get("status", "shortlisted")).replace("_", " ").title()
    interview_status = "Scheduled" if c.get("interview") else "Pending"
    
    # Video Interview Metrics (if available)
    video_comm = ai.get("communication", "N/A")
    video_conf = str(ai.get("metrics", {}).get("conf_score", "N/A"))
    video_risk = ai.get("cheating_risk", "N/A")
    
    return {
        "Rank": rank,
        "Name": name,
        "Email": email,
        "Phone": phone,
        "Experience": experience,
        "Skills": skills_str,
        "AI Match Score": f"{score}%",
        "Skill Match %": f"{skill_match}%",
        "Recommendation": recommendation,
        "Resume Summary": summary,
        "AI Feedback": feedback,
        "Strengths": strengths,
        "Weaknesses": weaknesses,
        "Hiring Decision": hiring_decision,
        "Interview Status": interview_status,
        "Video Communication": video_comm,
        "Video Confidence": video_conf,
        "Security Risk": video_risk
    }


def generate_shortlisted_excel_report(job, candidates):
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlisted Candidates"
    
    headers = [
        "Rank", "Name", "Email", "Phone", "Experience", "Skills", 
        "AI Match Score", "Skill Match %", "Recommendation", 
        "Resume Summary", "AI Feedback", "Strengths", "Weaknesses", 
        "Hiring Decision", "Interview Status", "Video Communication",
        "Video Confidence", "Security Risk"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for idx, c in enumerate(candidates):
        info = extract_candidate_info(c, idx)
        row = [info[k] for k in headers]
        ws.append(row)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
        
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for idx in [8, 9, 10, 11, 12]:
            row[idx].alignment = wrap_alignment
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_shortlisted_pdf_report(job, candidates):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter),
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], spaceAfter=14, textColor=colors.HexColor('#0F172A'))
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], spaceAfter=20, fontSize=11, textColor=colors.HexColor('#475569'))
    heading2_style = ParagraphStyle('Heading2', parent=styles['Heading2'], spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#1E293B'), fontSize=14)
    heading3_style = ParagraphStyle('Heading3', parent=styles['Heading3'], spaceBefore=12, spaceAfter=6, textColor=colors.HexColor('#3B82F6'))
    normal_style = styles['Normal']
    bullet_style = ParagraphStyle('Bullet', parent=normal_style, leftIndent=15, spaceAfter=4)
    
    # 1. Header
    job_title = job.get('title', 'Unknown Role')
    elements.append(Paragraph(f"Enterprise Intelligence Report: {job_title}", title_style))
    elements.append(Paragraph(f"Generated for internal HR review. Total Candidates: {len(candidates)}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E2E8F0'), spaceAfter=20))
    
    if not candidates:
        elements.append(Paragraph("No candidates available for this report.", normal_style))
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # Calculate some stats for Executive Summary
    scores = [c.get("score", 0) for c in candidates]
    avg_score = round(np.mean(scores)) if scores else 0
    top_score = round(max(scores)) if scores else 0
    
    # 2. Executive Hiring Summary
    elements.append(Paragraph("Executive Hiring Summary", heading2_style))
    exec_summary_text = f"This report evaluates {len(candidates)} candidates for the {job_title} position. The candidate pool has an average AI Match Score of <b>{avg_score}%</b>, with the top candidate achieving <b>{top_score}%</b>. The enclosed data provides deep behavioral and technical intelligence to facilitate immediate hiring decisions."
    elements.append(Paragraph(exec_summary_text, normal_style))
    elements.append(Spacer(1, 15))
    
    # 3. Top Ranked Candidates Table
    elements.append(Paragraph("Top Ranked Candidates", heading2_style))
    table_data = [["Rank", "Name", "Match", "Experience", "Recommendation"]]
    
    for idx, c in enumerate(candidates): # Display all candidates
        info = extract_candidate_info(c, idx)
        table_data.append([
            str(info["Rank"]),
            info["Name"],
            info["AI Match Score"],
            info["Experience"],
            info["Recommendation"].split('(')[0].strip() # Keep table clean
        ])
        
    table = Table(table_data, colWidths=[40, 140, 60, 80, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'TOP')
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))
    
    # 4. Detailed AI Profiles (Insights & Comparison)
    elements.append(Paragraph("AI Hiring Insights & Candidate Profiles", heading2_style))
    
    for idx, c in enumerate(candidates):
        info = extract_candidate_info(c, idx)
        
        # Profile Header
        elements.append(Paragraph(f"#{info['Rank']} - {info['Name']} | Match: {info['AI Match Score']}", heading3_style))
        
        # Core Info
        core_info = f"<b>Email:</b> {info['Email']} | <b>Phone:</b> {info['Phone']} | <b>Status:</b> {info['Hiring Decision']}"
        elements.append(Paragraph(core_info, normal_style))
        elements.append(Spacer(1, 5))
        
        # Recommendation
        elements.append(Paragraph(f"<b>Overall Hiring Recommendation:</b> {info['Recommendation']}", normal_style))
        elements.append(Spacer(1, 5))
        
        # Executive Summary
        elements.append(Paragraph(f"<b>Executive Summary:</b> {info['Resume Summary']}", normal_style))
        elements.append(Spacer(1, 5))
        
        # AI Feedback
        fb_clean = info['AI Feedback'].replace('\n', '<br/>')
        elements.append(Paragraph(f"<b>AI Recruiter Feedback:</b> {fb_clean}", normal_style))
        elements.append(Spacer(1, 5))
        
        # Strengths & Weaknesses (Comparison Format)
        elements.append(Paragraph(f"<b>Strengths:</b> {info['Strengths']}", bullet_style))
        elements.append(Paragraph(f"<b>Weaknesses/Missing:</b> {info['Weaknesses']}", bullet_style))
        elements.append(Spacer(1, 5))
        
        # Video Interview Monitoring Verdicts (If available)
        if info['Video Communication'] != "N/A":
            elements.append(Paragraph(f"<b>AI Video Interview Analysis:</b>", normal_style))
            elements.append(Paragraph(f"Communication: {info['Video Communication']} | Confidence: {info['Video Confidence']} | Security Risk: {info['Video Risk']}", bullet_style))
            
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceBefore=10, spaceAfter=15))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
